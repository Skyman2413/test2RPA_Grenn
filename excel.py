import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

path = "exchange.xlsx"
first_row = ["Дата", "Курс", "Изменение"]
wb = openpyxl.Workbook()
sheet_name = wb.sheetnames
sheet = wb[sheet_name[0]]
rows_cnt = 0

for cellObj in sheet['A1':'F1']:
    i = 0
    for cell in cellObj:
        cell.value = f'{first_row[i % 3]}'
        cell.alignment = Alignment(horizontal='center')
        i += 1


def optimize_columns_width():
    for i in range(1, 8):
        cell = sheet.cell(row=2, column=i)
        sheet.column_dimensions[get_column_letter(cell.column)] \
            .width = max(len(str(cell.value)), len(str(sheet[f'{get_column_letter(i)}1'].value))) + 6
    wb.save(path)


def add_to_excel(data, i):
    date_letter = get_column_letter(1 + i * 3)
    value_letter = get_column_letter(2 + i * 3)
    change_letter = get_column_letter(3 + i * 3)
    global rows_cnt
    rows_cnt = len(data) + 1
    for cellObj in sheet[f'{date_letter}2':f'{change_letter}{rows_cnt}']:
        for cell in cellObj:

            if cell.row == (len(data) + 1) and get_column_letter(cell.column) == change_letter:
                continue
            if get_column_letter(cell.column) == date_letter:
                cell.value = data[cell.row - 2][0]
                cell.number_format = 'DateTime'
            if get_column_letter(cell.column) == value_letter:
                cell.value = float(data[cell.row - 2][1])
                cell.style = 'Currency'
                if i == 0:
                    cell.number_format = '#,####0.0000$'
                else:
                    cell.number_format = '#,####0.0000€'

            if get_column_letter(cell.column) == change_letter:
                cell.value = f'={value_letter}{cell.row}-{value_letter}{cell.row + 1}'
            cell.alignment = Alignment(horizontal='center')


def fill_G_column():
    for cellObj in sheet[f'G2:G{rows_cnt}']:
        for cell in cellObj:
            cell.value = f'=E{cell.row}/B{cell.row}'
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = '#,####0.0000'
